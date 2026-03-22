import openpyxl
import json

# 读取excel数据并转换为json格式
workbook = openpyxl.load_workbook('source.xlsx')
worksheet = workbook.active

data = []
for row in worksheet.iter_rows():
    record = {}
    for cell in row:
        record[cell.column_letter] = cell.value
    data.append(record)


# 将数据写入另一个excel文件
new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active
new_worksheet.title = '数据源'

for record in data:
    row = []
    for key,value in record.items():
        if key == 'C':
            row.append(value)
    new_worksheet.append(row)

new_workbook.save('target.xlsx')

print("导出数据成功: target.xlsx")